'''
Created on 28 Oct 2020

@author: 612563313
'''

import openpyxl

def getRowCount(xwbFN,xws):
    XLwrkbook = openpyxl.load_workbook(xwbFN)
    XLwrkSheet = XLwrkbook[xws]
    return (XLwrkSheet.max_row)

def getColCount(xwbFN,xws):
    XLwrkbook = openpyxl.load_workbook(xwbFN)
    XLwrkSheet = XLwrkbook[xws]
    return (XLwrkSheet.max_column)


def getXLdata(xwbFN,xws,rownum,colnum):
    XLwrkbook = openpyxl.load_workbook(xwbFN)
    XLwrkSheet = XLwrkbook[xws]
    return XLwrkSheet.cell(rownum,colnum).value

def setXLdata(xwbFN,xws,rownum,colnum,data):
    XLwrkbook = openpyxl.load_workbook(xwbFN)
    XLwrkSheet = XLwrkbook[xws]
    XLwrkSheet.cell(rownum,colnum).value = data
    XLwrkbook.save(xwbFN)
    