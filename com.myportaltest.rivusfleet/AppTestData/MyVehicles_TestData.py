'''
Created on 28 Oct 2020

@author: 612563313
'''

from openpyxl import *  
from builtins import staticmethod

class MyVehicles_TestData:
   
#     Login Testdata
    LP_ssName = 'LoginPage'
    LP_title1 = 'Sign in to your account'
    LP_title2 = 'Sign In'
    
    
        #     MyVehiclSearch Testdata
    VP_ssName = 'MyVehicleSearch_'
    VP_title = 'My Vehicles'
    DwnLoad_fp = ".\\Reports\\AppRep_download\\"
    InRegNo = "WN15XYO"
    InOptVal = "Vehicle Documents"
     
    @staticmethod
    def MyVehicleDetails_td():
        Xwb = load_workbook(r'.\\ExternTestData\\XLdataToTup.xlsx')  
        Xws = Xwb["ParametrizeDataList"]
        rowCnt=  Xws.max_row
        DataLst = []
        for i in range(2,rowCnt+1):
            Col1 = Xws.cell(i,1).value
            Col2 = Xws.cell(i,2).value
            tup = (Col1,Col2)
            DataLst.append(tup)
        Xwb.close()    
        return(DataLst)
 
